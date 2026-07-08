"""
PaddleOCR Demo
──────────────
Supports: images (JPG, PNG, BMP, TIFF), scanned PDFs, and PDFs with an
existing embedded text layer (auto-detected — see extract_pdf_text_layer).
Languages: French (fr), Arabic (ar), English (en)

On first run, PaddleOCR downloads the model (~100 MB).
Subsequent runs use the cached model and are instant.

Dependencies:
    pip install paddleocr paddlepaddle pdf2image pillow openpyxl pymupdf
    # For PDF support also install poppler:
    # Windows: https://github.com/oschwartz10612/poppler-windows/releases
    # Then add poppler/Library/bin to your PATH
"""

import os
import sys
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading
import difflib
from typing import Dict, List, Tuple, Any, Optional
from html.parser import HTMLParser


# ─────────────────────────────────────────────
#  LAZY IMPORTS (heavy, load only when needed)
# ─────────────────────────────────────────────

def import_paddleocr():
    try:
        from paddleocr import PaddleOCR
        return PaddleOCR
    except ImportError:
        messagebox.showerror("Missing dependency",
            "PaddleOCR not installed.\n\nRun:\n  pip install paddleocr paddlepaddle")
        return None

def import_ppstructure():
    try:
        from paddleocr import PPStructure
        return PPStructure
    except ImportError:
        messagebox.showerror("Missing dependency",
            "PPStructure not installed/found.\n\nRun:\n  pip install paddleocr paddlepaddle")
        return None

def import_pdf2image():
    try:
        from pdf2image import convert_from_path
        return convert_from_path
    except ImportError:
        return None

def import_fitz():
    try:
        import fitz  # PyMuPDF
        return fitz
    except ImportError:
        return None

def import_pil():
    try:
        from PIL import Image, ImageTk, ImageDraw
        return Image, ImageTk, ImageDraw
    except ImportError:
        messagebox.showerror("Missing dependency",
            "Pillow not installed.\n\nRun:\n  pip install pillow")
        return None, None, None

def import_openpyxl():
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
        return openpyxl, get_column_letter
    except ImportError:
        messagebox.showerror("Missing dependency",
            "openpyxl not installed.\n\nRun:\n  pip install openpyxl")
        return None, None


# ─────────────────────────────────────────────
#  OCR LOGIC
# ─────────────────────────────────────────────

LANG_MAP = {
    "French":  "fr",
    "Arabic":  "ar",
    "English": "en",
}

_ocr_instances = {}

def get_ocr(lang_code: str):
    if lang_code not in _ocr_instances:
        PaddleOCR = import_paddleocr()
        if PaddleOCR is None:
            return None
        _ocr_instances[lang_code] = PaddleOCR(
            lang=lang_code,
            use_textline_orientation=False,
        )
    return _ocr_instances[lang_code]

# Also cache PPStructure engines
_table_instances = {}

def get_table_engine(lang_code: str):
    if lang_code not in _table_instances:
        PPStructure = import_ppstructure()
        if PPStructure is None:
            return None
        _table_instances[lang_code] = PPStructure(
            lang=lang_code,
            show_log=True,
            image_orientation=False
        )
    return _table_instances[lang_code]


def ocr_image(image_path: str, lang_code: str) -> List[Tuple[str, float, Any]]:
    ocr = get_ocr(lang_code)
    if ocr is None:
        return []
    lines = []
    results = ocr.predict(image_path)
    for res in results:
        # Depending on the API version, res format might change. This handles v3.x format
        texts  = res.get('rec_texts', [])
        scores = res.get('rec_scores', [])
        polys  = res.get('dt_polys', [None] * len(texts))
        for text, conf, bbox in zip(texts, scores, polys):
            if text and str(text).strip():
                lines.append((str(text), round(float(conf) * 100, 1), bbox))
    return lines


def ocr_pdf(pdf_path: str, lang_code: str, log_fn) -> Dict[int, List[Tuple[str, float, Any]]]:
    convert_from_path = import_pdf2image()
    if convert_from_path is None:
        messagebox.showerror("Missing dependency",
            "pdf2image not installed.\n\nRun:\n  pip install pdf2image\n\n"
            "Also install Poppler and add it to PATH.")
        return {}

    log_fn("Converting PDF pages to images...")
    try:
        pages = convert_from_path(pdf_path, dpi=200)
    except Exception as e:
        messagebox.showerror("PDF Error", f"Could not convert PDF:\n{e}\n\n"
                             "Make sure Poppler is installed and in PATH.")
        return {}

    results = {}
    for i, page_img in enumerate(pages, 1):
        log_fn(f"OCR — page {i}/{len(pages)}...")
        tmp = pdf_path + f"_page{i}.png"
        page_img.save(tmp)
        try:
            results[i] = ocr_image(tmp, lang_code)
        finally:
            if os.path.exists(tmp):
                os.remove(tmp)
    return results


# ─────────────────────────────────────────────
#  PDF TEXT-LAYER DETECTION (skip OCR when the PDF already has
#  trustworthy embedded text — verified against an OCR sample, not
#  just "does copiable text exist", since some invoicing software
#  produces a text layer with broken accented-character mappings)
# ─────────────────────────────────────────────

MIN_TEXT_LAYER_CHARS = 50       # below this, treat as "no usable text layer"
TEXT_LAYER_TRUST_THRESHOLD = 0.82  # similarity vs OCR probe needed to trust the text layer


def extract_pdf_text_layer(pdf_path: str) -> Dict[int, List[Tuple[str, float, Any]]]:
    """Extracts text directly from a PDF's embedded text layer via PyMuPDF —
    no rasterization, no OCR. Returns the same {page: [(text, confidence, bbox), ...]}
    shape as ocr_pdf/ocr_image, with confidence fixed at 100.0 since it's a
    direct read, not a model prediction."""
    fitz = import_fitz()
    if fitz is None:
        return {}
    results: Dict[int, List[Tuple[str, float, Any]]] = {}
    try:
        doc = fitz.open(pdf_path)
        for i, page in enumerate(doc, 1):
            lines = []
            d = page.get_text("dict")
            for block in d.get('blocks', []):
                for line in block.get('lines', []):
                    text = "".join(span.get('text', '') for span in line.get('spans', [])).strip()
                    if text:
                        lines.append((text, 100.0, line.get('bbox')))
            results[i] = lines
        doc.close()
    except Exception:
        return {}
    return results


def ocr_pdf_page(pdf_path: str, lang_code: str, page_num: int) -> List[Tuple[str, float, Any]]:
    """OCRs a single PDF page (used as a quick probe to sanity-check the text layer)."""
    convert_from_path = import_pdf2image()
    if convert_from_path is None:
        return []
    try:
        pages = convert_from_path(pdf_path, dpi=200, first_page=page_num, last_page=page_num)
    except Exception:
        return []
    if not pages:
        return []
    tmp = pdf_path + f"_probe_page{page_num}.png"
    pages[0].save(tmp)
    try:
        return ocr_image(tmp, lang_code)
    finally:
        if os.path.exists(tmp):
            os.remove(tmp)


def text_layer_trust_score(layer_text: str, probe_text: str) -> float:
    """Similarity (0-1) between the PDF's embedded text on page 1 and a quick
    OCR pass on the same page. Used to catch text layers that exist but are
    corrupted (e.g. broken font encoding dropping accented characters).
    Not perfect — the OCR probe itself isn't 100% accurate either — so treat
    TEXT_LAYER_TRUST_THRESHOLD as a tunable heuristic, not ground truth."""
    a = layer_text.lower().strip()
    b = probe_text.lower().strip()
    if not a or not b:
        return 0.0
    return difflib.SequenceMatcher(None, a, b).ratio()


def format_results_text(lines: List[Tuple[str, float, Any]], show_confidence: bool=True) -> str:
    if not lines:
        return "(no text detected)"
    parts = []
    for text, conf, _ in lines:
        if show_confidence:
            parts.append(f"{text}  [{conf}%]")
        else:
            parts.append(text)
    return "\n".join(parts)


# ─────────────────────────────────────────────
#  MARKDOWN EXPORT (built from PPStructure layout regions)
# ─────────────────────────────────────────────

class _MDTableParser(HTMLParser):
    """Parses a PPStructure table HTML string into rows, tracking which rows
    contained <th> cells so we know which row to use as the header."""
    def __init__(self):
        super().__init__()
        self.rows: List[List[str]] = []
        self.header_flags: List[bool] = []
        self._current_row: List[str] = []
        self._row_has_th = False
        self._in_cell = False
        self._cell_data = ""

    def handle_starttag(self, tag, attrs):
        if tag in ('td', 'th'):
            self._in_cell = True
            self._cell_data = ""
            if tag == 'th':
                self._row_has_th = True
        elif tag == 'tr':
            self._current_row = []
            self._row_has_th = False

    def handle_endtag(self, tag):
        if tag in ('td', 'th'):
            self._in_cell = False
            self._current_row.append(self._cell_data.strip())
        elif tag == 'tr':
            if self._current_row:
                self.rows.append(self._current_row)
                self.header_flags.append(self._row_has_th)
            self._current_row = []

    def handle_data(self, data):
        if self._in_cell:
            self._cell_data += data


def _md_escape_cell(text: str) -> str:
    return text.replace("|", "\\|").replace("\n", "<br>").strip()


def html_table_to_markdown(html_str: str) -> str:
    """Converts a PPStructure table HTML string into a GFM markdown table.

    Known limitation: this does not merge rowspan/colspan cells, so tables
    with merged cells will show repeated or empty cells instead of a true
    merge. Fine for simple/regular tables; complex merged tables may need
    a manual pass.
    """
    if not html_str:
        return ""
    parser = _MDTableParser()
    parser.feed(html_str)
    rows = parser.rows
    if not rows:
        return ""

    n_cols = max(len(r) for r in rows)

    def pad(row):
        return row + [""] * (n_cols - len(row))

    header_idx = next((i for i, f in enumerate(parser.header_flags) if f), 0)
    header = pad(rows[header_idx])
    body = [pad(r) for i, r in enumerate(rows) if i != header_idx]

    lines = ["| " + " | ".join(_md_escape_cell(c) for c in header) + " |",
             "| " + " | ".join(["---"] * n_cols) + " |"]
    for r in body:
        lines.append("| " + " | ".join(_md_escape_cell(c) for c in r) + " |")
    return "\n".join(lines)


def _region_lines(res: Any) -> List[str]:
    """Extracts a flat list of text lines from a PPStructure region's `res` field."""
    lines = []
    if isinstance(res, list):
        for item in res:
            if isinstance(item, dict) and 'text' in item and str(item['text']).strip():
                lines.append(str(item['text']).strip())
    elif isinstance(res, dict) and 'text' in res:
        text = str(res['text']).strip()
        if text:
            lines.append(text)
    elif isinstance(res, str) and res.strip():
        lines.append(res.strip())
    return lines


def _region_line_height(region: Dict) -> Optional[float]:
    """Rough proxy for font size: region bbox height / its line count.
    Used only to compare relative sizes within the same document, not as
    an absolute measurement."""
    bbox = region.get('bbox')
    lines = _region_lines(region.get('res', {}))
    if not bbox or len(bbox) != 4 or not lines:
        return None
    height = bbox[3] - bbox[1]
    if height <= 0:
        return None
    return height / len(lines)


def _region_sort_key(region: Dict) -> Tuple[float, float]:
    bbox = region.get('bbox') or [0, 0, 0, 0]
    return (bbox[1], bbox[0])  # top-to-bottom, then left-to-right


def build_markdown_document(results: Dict[int, List[Dict]]) -> str:
    """Builds a structured Markdown document from PPStructure layout results.

    Heading detection is heuristic, not exact: a text region is promoted to
    a heading if it's short (<=2 lines, <=12 words) AND its estimated line
    height is noticeably taller than the document's median body-text line
    height, or if PPStructure already tagged the region as 'title'.
    Thresholds (1.3x / 1.8x median) are conservative starting points —
    tune them if headings are over/under-triggering on your documents.
    """
    samples = []
    for page_regions in results.values():
        for region in page_regions:
            if region.get('type') not in ('table', 'figure'):
                h = _region_line_height(region)
                if h:
                    samples.append(h)
    baseline = sorted(samples)[len(samples) // 2] if samples else None  # median

    page_blocks = []
    for page_num in sorted(results.keys()):
        regions = sorted(results[page_num], key=_region_sort_key)
        blocks = []
        for region in regions:
            r_type = region.get('type', 'text')
            res = region.get('res', {})

            if r_type == 'table':
                html_str = res.get('html', '') if isinstance(res, dict) else (res if isinstance(res, str) else '')
                md_table = html_table_to_markdown(html_str)
                if md_table:
                    blocks.append(md_table)
                continue

            if r_type == 'figure':
                blocks.append("*(figure/image — not extracted)*")
                continue

            lines = _region_lines(res)
            if not lines:
                continue

            if r_type == 'list':
                blocks.append("\n".join(f"- {ln}" for ln in lines))
                continue

            text = " ".join(lines)
            word_count = len(text.split())
            line_height = _region_line_height(region)

            is_heading = False
            level = 3
            if r_type == 'title':
                is_heading, level = True, 2
            if baseline and line_height and len(lines) <= 2 and word_count <= 12:
                ratio = line_height / baseline
                if ratio >= 1.8:
                    is_heading, level = True, 1
                elif ratio >= 1.3:
                    is_heading, level = True, 2

            blocks.append(f"{'#' * level} {text}" if is_heading else text)

        page_blocks.append("\n\n".join(blocks))

    if len(page_blocks) > 1:
        return "\n\n<!-- page break -->\n\n".join(page_blocks)
    return page_blocks[0] if page_blocks else ""


# ─────────────────────────────────────────────
#  GUI
# ─────────────────────────────────────────────

class PaddleOCRDemo:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("PaddleOCR Demo")
        self.root.geometry("1020x720")
        self.root.resizable(True, True)

        self.current_file: Optional[str] = None
        self.ocr_results: Dict[int, List[Tuple[str, float, Any]]] = {}
        self.current_page: int = 1
        
        # TK Variables explicitly typed
        self.lang_var = tk.StringVar(value="French")
        self.show_conf = tk.BooleanVar(value=True)
        self.page_var = tk.StringVar(value="—")
        self.status_var = tk.StringVar(value="Ready — open an image or PDF then click Run OCR.")
        
        # UI Elements
        self.file_label: tk.Label
        self.preview_label: tk.Label
        self.text_out: tk.Text
        self.progress: ttk.Progressbar
        
        # Table extractions state
        # format: page_num -> list of regions dict: {'type': 'table'|'text', 'bbox': [x1,y1,x2,y2], 'res': ...}
        self.table_results: Dict[int, List[Dict]] = {}
        self.preview_mode = 'ocr' # 'ocr' or 'table'

        self._build_ui()

    def _build_ui(self):
        # ── Top bar ────────────────────────────────────────────
        top = tk.Frame(self.root, pady=8, padx=12)
        top.pack(fill='x')

        tk.Label(top, text="PaddleOCR Demo", font=("Arial", 14, "bold")).pack(side='left')

        tk.Label(top, text="Language:", font=("Arial", 10)).pack(side='left', padx=(20, 4))
        ttk.Combobox(top, textvariable=self.lang_var,
                     values=list(LANG_MAP.keys()), width=10, state='readonly').pack(side='left')

        tk.Checkbutton(top, text="Show confidence %", variable=self.show_conf,
                       command=self._refresh_text).pack(side='left', padx=(16, 0))

        # ── Buttons ────────────────────────────────────────────
        btn = tk.Frame(self.root, padx=12, pady=4)
        btn.pack(fill='x')

        for label, cmd, color in [
            ("📂  Open Image",  self._open_image,   "#2E75B6"),
            ("📄  Open PDF",    self._open_pdf,     "#2E75B6"),
            ("📁  Batch Folder", self._batch_folder, "#2E75B6"),
            ("▶  Run OCR",      self._run_ocr,      "#375623"),
            ("🗂  Extract Tables",self._extract_tables,"#7B2D8E"),
            ("📝  Export Markdown", self._export_markdown, "#B45309"),
            ("💾  Save Text",   self._save_text,    "#595959"),
        ]:
            if "Batch" in label or "Extract" in label or "Export" in label:
                 tk.Button(btn, text=label, command=cmd, width=16,
                      bg=color, fg="white",
                      font=("Arial", 10, "bold")).pack(side='left', padx=(12, 4))
            else:
                tk.Button(btn, text=label, command=cmd, width=16,
                        bg=color, fg="white",
                        font=("Arial", 10, "bold")).pack(side='left', padx=4)

        # ── File label ─────────────────────────────────────────
        self.file_label = tk.Label(self.root, text="No file selected.",
                                   font=("Arial", 9), fg="#555", anchor='w')
        self.file_label.pack(fill='x', padx=14)

        # ── Split pane ─────────────────────────────────────────
        pane = tk.PanedWindow(self.root, orient='horizontal', sashwidth=6, bg="#bbb")
        pane.pack(fill='both', expand=True, padx=8, pady=6)

        # Left: preview
        left = tk.Frame(pane, bg="#f0f0f0")
        pane.add(left, minsize=320)

        tk.Label(left, text="Preview", font=("Arial", 9, "bold"),
                 bg="#f0f0f0").pack(anchor='w', padx=6, pady=(4, 0))

        self.preview_label = tk.Label(left, bg="#d8d8d8", text="No image loaded",
                                      relief='sunken')
        self.preview_label.pack(fill='both', expand=True, padx=6, pady=(4, 2))

        nav = tk.Frame(left, bg="#f0f0f0")
        nav.pack(pady=(0, 6))
        tk.Button(nav, text="◀", command=self._prev_page, width=4).pack(side='left')
        tk.Label(nav, textvariable=self.page_var, width=12,
                 font=("Arial", 9), bg="#f0f0f0").pack(side='left')
        tk.Button(nav, text="▶", command=self._next_page, width=4).pack(side='left')

        # Right: text output
        right = tk.Frame(pane)
        pane.add(right, minsize=380)

        tk.Label(right, text="OCR Output", font=("Arial", 9, "bold")).pack(anchor='w', padx=6, pady=(4, 0))

        tf = tk.Frame(right)
        tf.pack(fill='both', expand=True, padx=6, pady=6)

        self.text_out = tk.Text(tf, wrap='word', font=("Consolas", 10),
                                relief='sunken', borderwidth=1, state='normal')
        sb = ttk.Scrollbar(tf, command=self.text_out.yview)
        self.text_out.configure(yscrollcommand=sb.set)
        sb.pack(side='right', fill='y')
        self.text_out.pack(side='left', fill='both', expand=True)

        # ── Progress + status ──────────────────────────────────
        self.progress = ttk.Progressbar(self.root, mode='indeterminate')
        # (packed/unpacked dynamically)

        tk.Label(self.root, textvariable=self.status_var, anchor='w',
                 font=("Arial", 9), fg="#333", bg="#e8e8e8",
                 relief='sunken').pack(fill='x', side='bottom')

    # ── File opening ───────────────────────────────────────────

    def _open_image(self):
        path = filedialog.askopenfilename(
            title="Select image",
            filetypes=[("Images", "*.jpg *.jpeg *.png *.bmp *.tiff *.tif"),
                       ("All files", "*.*")]
        )
        if not path:
            return
        self.current_file = path
        self.ocr_results  = {}
        self.table_results = {}
        self.current_page = 1
        self.preview_mode = 'ocr'
        self.file_label.config(text=f"File: {path}")
        self._show_image_preview(path)
        self._set_text("Image loaded. Click ▶ Run OCR or 🗂 Extract Tables.")
        self.status_var.set(f"Loaded: {os.path.basename(path)}")
        self._update_page_nav()

    def _open_pdf(self):
        path = filedialog.askopenfilename(
            title="Select PDF",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
        )
        if not path:
            return
        self.current_file = path
        self.ocr_results  = {}
        self.table_results = {}
        self.current_page = 1
        self.preview_mode = 'ocr'
        self.file_label.config(text=f"File: {path}")
        self.preview_label.config(image='', text="PDF loaded\n(preview not available\nbefore extraction)")
        self._set_text("PDF loaded. Checking for an embedded text layer...")
        self.status_var.set(f"Loaded: {os.path.basename(path)} — checking text layer...")
        self._update_page_nav()
        self._auto_detect_text_layer()

    def _auto_detect_text_layer(self):
        """On PDF load: try the embedded text layer first, verify it against a
        quick OCR probe of page 1, and only fall back to full OCR if the text
        layer is missing or looks unreliable. Saves OCR time on born-digital
        or already-OCR'd PDFs; protects against PDFs whose embedded text is
        present but corrupted (broken font encoding)."""
        self.progress.configure(mode='indeterminate')
        self.progress.pack(fill='x', padx=8, side='bottom', before=self.root.winfo_children()[-1])
        self.progress.start(12)
        self.root.update()

        def worker():
            if self.current_file is None:
                return
            layer_results = extract_pdf_text_layer(self.current_file)
            total_chars = sum(len(t) for lines in layer_results.values() for t, _, _ in lines)

            if total_chars < MIN_TEXT_LAYER_CHARS:
                # No usable embedded text — clearly needs OCR, skip the probe
                self.root.after(0, lambda: self._on_text_layer_check_done(layer_results, None))
                return

            self._log("Text layer found — verifying against a quick OCR sample...")
            lang_code = LANG_MAP[self.lang_var.get()]
            try:
                probe_lines = ocr_pdf_page(self.current_file, lang_code, page_num=1)
            except Exception:
                probe_lines = []

            self.root.after(0, lambda: self._on_text_layer_check_done(layer_results, probe_lines))

        threading.Thread(target=worker, daemon=True).start()

    def _on_text_layer_check_done(self, layer_results: Dict[int, List[Tuple]], probe_lines: Optional[List[Tuple]]):
        self.progress.stop()
        self.progress.pack_forget()

        if not layer_results or not any(layer_results.values()):
            self._set_text("No embedded text layer detected — this looks like a scanned PDF.\n\nClick ▶ Run OCR.")
            self.status_var.set("No text layer found — ready for OCR.")
            return

        if not probe_lines:
            # Couldn't run a probe (e.g. pdf2image/Poppler missing) — use the
            # text layer but flag that it wasn't verified.
            self.ocr_results = layer_results
            self.current_page = 1
            self._refresh_text()
            self._update_page_nav()
            self.status_var.set(
                "Text layer loaded directly (unverified — could not run the OCR sanity check). "
                "Click ▶ Run OCR to force real OCR instead."
            )
            return

        probe_text = " ".join(t for t, _, _ in probe_lines)
        layer_page1_text = " ".join(t for t, _, _ in layer_results.get(1, []))
        score = text_layer_trust_score(layer_page1_text, probe_text)

        if score >= TEXT_LAYER_TRUST_THRESHOLD:
            self.ocr_results = layer_results
            self.current_page = 1
            self._refresh_text()
            self._update_page_nav()
            self.status_var.set(
                f"Text layer used directly (match {score*100:.0f}% vs OCR sample) — no OCR needed. "
                "Click ▶ Run OCR to force it anyway."
            )
        else:
            self.status_var.set(f"Text layer looks unreliable (match {score*100:.0f}% vs OCR sample) — running full OCR...")
            self._set_text(
                f"Embedded text layer looks corrupted (only {score*100:.0f}% match against a quick "
                "OCR check on page 1) — likely a font-encoding issue in the source PDF.\n\n"
                "Falling back to full OCR..."
            )
            self._run_ocr()

    # ── OCR ────────────────────────────────────────────────────

    def _run_ocr(self):
        if not self.current_file:
            messagebox.showwarning("No file", "Please open an image or PDF first.")
            return

        lang_code = LANG_MAP[self.lang_var.get()]
        self._set_text(
            f"Running OCR in {self.lang_var.get()}...\n\n"
            "⏳ First run downloads the model (~100 MB) — this may take a minute.\n"
            "Subsequent runs are instant (model cached locally)."
        )
        self.status_var.set("Running OCR...")
        self.preview_mode = 'ocr'
        self.progress.configure(mode='indeterminate')
        self.progress.pack(fill='x', padx=8, side='bottom', before=self.root.winfo_children()[-1])
        self.progress.start(12)
        self.root.update()

        def worker():
            if self.current_file is None:
                return
            try:
                ext = os.path.splitext(self.current_file)[1].lower()
                if ext == '.pdf':
                    results = ocr_pdf(self.current_file, lang_code, self._log)
                else:
                    lines   = ocr_image(self.current_file, lang_code)
                    results = {1: lines}
                self.ocr_results  = results
                self.current_page = 1
                self.root.after(0, self._on_ocr_done)
            except Exception as e:
                self.root.after(0, lambda err=str(e): self._on_ocr_error(err))

        threading.Thread(target=worker, daemon=True).start()

    def _on_ocr_done(self):
        self.progress.stop()
        self.progress.pack_forget()
        total = sum(len(v) for v in self.ocr_results.values())
        self.status_var.set(
            f"Done — {len(self.ocr_results)} page(s), {total} line(s) detected."
        )
        self._refresh_text()
        self._update_page_nav()
        
        # Display preview for single or multi-page PDF (using generated PNG)
        self._refresh_preview()

    def _on_ocr_error(self, err: str):
        self.progress.stop()
        self.progress.pack_forget()
        self.status_var.set(f"Error: {err}")
        messagebox.showerror("OCR Error", f"An error occurred:\n\n{err}")

    # ── Batch Folder Mode ────────────────────────────────────────
    
    def _batch_folder(self):
        in_dir = filedialog.askdirectory(title="1/2: Select Input Folder (with images/PDFs)")
        if not in_dir:
            return
            
        out_dir = filedialog.askdirectory(title="2/2: Select Output Folder (for text & excel files)")
        if not out_dir:
            return
            
        # Collect files
        valid_exts = {'.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.tif', '.pdf'}
        target_files = []
        for root_dir, dirs, files in os.walk(in_dir):
            for file in files:
                ext = os.path.splitext(file)[1].lower()
                if ext in valid_exts:
                    target_files.append(os.path.join(root_dir, file))
                    
        if not target_files:
            messagebox.showinfo("Batch OCR", f"No images or PDFs found in:\n{in_dir}")
            return
            
        if not messagebox.askyesno("Batch OCR", f"Found {len(target_files)} files to process.\n"
                                   f"Language: {self.lang_var.get()}.\n"
                                   f"Run batch OCR now?"):
            return
            
        lang_code = LANG_MAP[self.lang_var.get()]
        
        # Setup UI
        self._set_text(f"Starting batch Processing of {len(target_files)} files...")
        self.progress.configure(mode='determinate', maximum=len(target_files), value=0)
        self.progress.pack(fill='x', padx=8, side='bottom', before=self.root.winfo_children()[-1])
        
        def worker():
            openpyxl_mod, get_col = import_openpyxl()
            if openpyxl_mod is None:
                self.root.after(0, lambda: self._on_ocr_error("openpyxl required for batch summary.\npip install openpyxl"))
                return
                
            wb = openpyxl_mod.Workbook()
            ws = wb.active
            ws.title = "Batch Summary"
            # Header
            headers = ["Filename", "Page", "Line", "Confidence %", "Text"]
            for col, h in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=h)
            current_row = 2
            
            successful = 0
            
            for i, filepath in enumerate(target_files):
                filename = os.path.basename(filepath)
                self._log(f"Batch ({i+1}/{len(target_files)}): {filename}")
                self.root.after(0, lambda val=i: self.progress.configure(value=val))
                
                try:
                    ext = os.path.splitext(filepath)[1].lower()
                    if ext == '.pdf':
                        results = ocr_pdf(filepath, lang_code, lambda msg: None) # suppress internal OCR logging
                    else:
                        lines = ocr_image(filepath, lang_code)
                        results = {1: lines}
                        
                    # Save local txt
                    txt_path = os.path.join(out_dir, f"{os.path.splitext(filename)[0]}_ocr.txt")
                    with open(txt_path, 'w', encoding='utf-8') as f:
                        for page_num in sorted(results.keys()):
                            page_lines = results[page_num]
                            if len(results) > 1:
                                f.write(f"\n{'='*40}\n  PAGE {page_num}\n{'='*40}\n\n")
                            f.write(format_results_text(page_lines, show_confidence=False))
                            f.write("\n")
                            
                            # Add to Excel
                            for line_idx, (text, conf, _) in enumerate(page_lines, 1):
                                ws.cell(row=current_row, column=1, value=filename)
                                ws.cell(row=current_row, column=2, value=page_num)
                                ws.cell(row=current_row, column=3, value=line_idx)
                                ws.cell(row=current_row, column=4, value=conf)
                                ws.cell(row=current_row, column=5, value=text)
                                current_row += 1
                                
                    successful += 1
                except Exception as e:
                    print(f"Error processing {filename}: {e}", file=sys.stderr)
                    # Write error to excel for visibility
                    ws.cell(row=current_row, column=1, value=filename)
                    ws.cell(row=current_row, column=5, value=f"ERROR: {str(e)}")
                    current_row += 1
                    
            try:
                summary_file = os.path.join(out_dir, "batch_summary.xlsx")
                wb.save(summary_file)
            except Exception as e:
                self.root.after(0, lambda err=str(e): self._on_ocr_error(f"Failed to save summary Excel: {err}"))
                return
                
            self.root.after(0, lambda: self._on_batch_done(successful, len(target_files), out_dir))
            
        threading.Thread(target=worker, daemon=True).start()

    def _on_batch_done(self, successful: int, total: int, out_dir: str):
        self.progress.pack_forget()
        self.status_var.set(f"Batch completed: {successful}/{total} successful.")
        self._set_text(f"Batch completed!\nProcessed {successful} out of {total} files successfully.\n\n"
                       f"Outputs saved to:\n{out_dir}\n"
                       f"↳ Individual .txt files\n"
                       f"↳ batch_summary.xlsx")
        messagebox.showinfo("Batch Complete", f"Processed {successful}/{total} files.\nSaved in {out_dir}")
        
    # ── Table Extraction ───────────────────────────────────────
    
    def _extract_tables(self):
        if not self.current_file:
            messagebox.showwarning("No file", "Please open an image or PDF first.")
            return
            
        lang_code = LANG_MAP[self.lang_var.get()]
        self._set_text(
            f"Extracting tables in {self.lang_var.get()}...\n\n"
            "This uses the PPStructure pipeline."
        )
        self.status_var.set("Extracting tables/layout...")
        self.preview_mode = 'table'
        self.progress.configure(mode='indeterminate')
        self.progress.pack(fill='x', padx=8, side='bottom', before=self.root.winfo_children()[-1])
        self.progress.start(12)
        self.root.update()

        def worker():
            if self.current_file is None:
                return
            openpyxl_mod, get_col = import_openpyxl()
            if openpyxl_mod is None:
                self.root.after(0, lambda: self._on_ocr_error("openpyxl required for table extraction.\npip install openpyxl"))
                return
                
            try:
                engine = get_table_engine(lang_code)
                if engine is None:
                    self.root.after(0, lambda: self._on_ocr_error("Failed to load PPStructure engine."))
                    return
                
                ext = os.path.splitext(self.current_file)[1].lower()
                results = {}
                
                if ext == '.pdf':
                    convert_from_path = import_pdf2image()
                    if convert_from_path is None:
                        self.root.after(0, lambda: self._on_ocr_error("pdf2image required for PDF table extraction."))
                        return
                    self._log("Converting PDF to images for table extraction...")
                    pages = convert_from_path(self.current_file, dpi=200)
                    for i, page_img in enumerate(pages, 1):
                        self._log(f"Extracting table — page {i}/{len(pages)}...")
                        tmp = self.current_file + f"_page{i}.png"
                        page_img.save(tmp)
                        try:
                            res = engine(tmp)
                            results[i] = res
                        finally:
                            if os.path.exists(tmp):
                                os.remove(tmp)
                else:
                    self._log(f"Extracting table...")
                    res = engine(self.current_file)
                    results[1] = res
                    
                self.table_results = results
                self.current_page = 1
                self.root.after(0, lambda: self._on_table_done(results))
            except Exception as e:
                self.root.after(0, lambda err=str(e): self._on_ocr_error(err))

        threading.Thread(target=worker, daemon=True).start()

    def _on_table_done(self, results: Dict[int, List[Dict]]):
        self.progress.stop()
        self.progress.pack_forget()
        
        # Count tables
        total_tables = 0
        for page_res in results.values():
            for region in page_res:
                if isinstance(region, dict) and region.get('type') == 'table':
                    total_tables += 1
                        
        self.status_var.set(
            f"Table extraction done — {len(results)} page(s), {total_tables} table(s) found."
        )
        self._refresh_table_text()
        self._update_page_nav()
        self._refresh_preview()
        
        if total_tables > 0:
            if messagebox.askyesno("Save Tables", f"Found {total_tables} tables.\nDo you want to export to Excel (.xlsx)?"):
                self._export_table_xlsx(results)
    
    def _export_table_xlsx(self, results: Dict[int, List[Dict]]):
        openpyxl_mod, get_col = import_openpyxl()
        if not openpyxl_mod: return
        
        default = os.path.splitext(os.path.basename(self.current_file))[0] + "_tables.xlsx"
        out_path = filedialog.asksaveasfilename(
            title="Save Extracted Tables",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=default
        )
        if not out_path: return
        
        try:
            wb = openpyxl_mod.Workbook()
            # Remove default sheet
            wb.remove(wb.active) 
            
            table_idx = 1
            has_text = False
            
            # Text gathering
            text_sheet = None
            text_row = 1
            
            # Simple HTML table parser (PPStructure returns dicts with 'html' inside 'res')
            from html.parser import HTMLParser
            class TableParser(HTMLParser):
                def __init__(self):
                    super().__init__()
                    self.rows = []
                    self.current_row = []
                    self.in_td = False
                    self.cell_data = ""
                def handle_starttag(self, tag, attrs):
                    if tag in ('td', 'th'):
                        self.in_td = True
                        self.cell_data = ""
                    elif tag == 'tr':
                        self.current_row = []
                def handle_endtag(self, tag):
                    if tag in ('td', 'th'):
                        self.in_td = False
                        self.current_row.append(self.cell_data.strip())
                    elif tag == 'tr':
                        self.rows.append(self.current_row)
                        self.current_row = []
                def handle_data(self, data):
                    if self.in_td:
                        self.cell_data += data
            
            for page_num in sorted(results.keys()):
                for region in results[page_num]:
                    r_type = region.get('type')
                    r_res = region.get('res', {})
                    
                    if r_type == 'table':
                        html_str = ""
                        if isinstance(r_res, dict) and 'html' in r_res:
                            html_str = r_res['html']
                        elif isinstance(r_res, list) and len(r_res)>0 and isinstance(r_res[0], dict) and 'text' in r_res[0]:
                            # Fallback if structure is different
                            html_str = "<table><tr><td>" + "</td></tr><tr><td>".join([x.get('text','') for x in r_res]) + "</td></tr></table>"
                        elif isinstance(r_res, str):
                            html_str = r_res
                            
                        # Parse HTML
                        parser = TableParser()
                        parser.feed(html_str)
                        
                        sheet = wb.create_sheet(f"Table_{table_idx}")
                        for row_idx, r_data in enumerate(parser.rows, 1):
                            for col_idx, cell_value in enumerate(r_data, 1):
                                sheet.cell(row=row_idx, column=col_idx, value=cell_value)
                        
                        table_idx += 1
                        
                    else:
                        # Non-table text
                        if not text_sheet:
                            text_sheet = wb.create_sheet("Text")
                            has_text = True
                            
                        text_val = ""
                        if isinstance(r_res, list):
                            for line in r_res:
                                if isinstance(line, dict) and 'text' in line:
                                    text_val += line['text'] + "\n"
                        elif isinstance(r_res, dict) and 'text' in r_res:
                            text_val = r_res['text']
                        elif isinstance(r_res, str):
                            text_val = r_res
                            
                        if text_val:
                            text_sheet.cell(row=text_row, column=1, value=f"[Page {page_num}]")
                            text_sheet.cell(row=text_row, column=2, value=text_val.strip())
                            text_row += 1
                            
            if len(wb.sheetnames) == 0:
                wb.create_sheet("Empty")
                
            wb.save(out_path)
            messagebox.showinfo("Saved", f"Tables exported to:\n{out_path}")
            
        except Exception as e:
             messagebox.showerror("Export Error", f"Failed to export tables to Excel:\n{e}")

    def _export_markdown(self):
        if not self.table_results:
            messagebox.showwarning(
                "No layout data",
                "Run 🗂 Extract Tables first — Markdown export uses the layout "
                "regions it detects (tables, titles, text blocks), not the "
                "plain ▶ Run OCR results."
            )
            return

        default = os.path.splitext(os.path.basename(self.current_file))[0] + ".md"
        out_path = filedialog.asksaveasfilename(
            title="Save Markdown",
            defaultextension=".md",
            filetypes=[("Markdown files", "*.md"), ("All files", "*.*")],
            initialfile=default
        )
        if not out_path:
            return

        try:
            md = build_markdown_document(self.table_results)
            with open(out_path, 'w', encoding='utf-8') as f:
                f.write(md)
            self.status_var.set(f"Saved: {out_path}")
            messagebox.showinfo("Saved", f"Markdown saved to:\n{out_path}")
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export Markdown:\n{e}")

    def _refresh_table_text(self):
        if not self.table_results:
            return
        page_res = self.table_results.get(self.current_page, [])
        parts = []
        for i, region in enumerate(page_res):
            r_type = region.get('type', 'unknown')
            box = region.get('bbox', [])
            parts.append(f"Region {i+1} [{r_type.upper()}]: BoundingBox {box}")
            
            res = region.get('res', {})
            if r_type == 'table':
                if isinstance(res, dict) and 'html' in res:
                    parts.append("  Contains HTML Table representation.")
                else:
                    parts.append("  Table detected (format varies).")
            else:
                text_val = ""
                if isinstance(res, list):
                     for line in res:
                          if isinstance(line, dict) and 'text' in line:
                              text_val += "  " + line['text'] + "\n"
                elif isinstance(res, dict) and 'text' in res:
                    text_val = "  " + res['text']
                if text_val:
                    parts.append(text_val.strip())
            parts.append("")
                
        txt = "\n".join(parts)
        if len(self.table_results) > 1:
            txt = f"── Page {self.current_page} / {len(self.table_results)} ──\n\n" + txt
        self._set_text(txt)

    # ── Helpers ────────────────────────────────────────────────

    def _log(self, msg: str):
        self.root.after(0, lambda m=msg: self.status_var.set(m))

    def _set_text(self, txt: str):
        self.text_out.delete('1.0', 'end')
        self.text_out.insert('end', txt)

    def _refresh_text(self):
        if self.preview_mode == 'table':
            self._refresh_table_text()
            return
            
        if not self.ocr_results:
            return
        lines = self.ocr_results.get(self.current_page, [])
        txt   = format_results_text(lines, self.show_conf.get())
        if len(self.ocr_results) > 1:
            txt = f"── Page {self.current_page} / {len(self.ocr_results)} ──\n\n" + txt
        self._set_text(txt)

    def _get_page_image_path(self) -> Optional[str]:
        if not self.current_file: return None
        ext = os.path.splitext(self.current_file)[1].lower()
        if ext != '.pdf':
            return self.current_file
        # For PDFs, we saved intermediate images iff we ran OCR/Table.
        # But those are deleted. We need to re-extract or fallback.
        # For simplicity in this demo, we'll try to re-extract just this page if requested.
        converted = self.current_file + f"_page{self.current_page}.png"
        if os.path.exists(converted):
            return converted
            
        # If it doesn't exist, we must re-generate it for preview
        convert_from_path = import_pdf2image()
        if convert_from_path:
            try:
                # Convert only the specific page (1-indexed)
                pages = convert_from_path(self.current_file, dpi=100, first_page=self.current_page, last_page=self.current_page)
                if pages:
                    pages[0].save(converted)
                    return converted
            except:
                pass
        return None

    def _refresh_preview(self):
        img_path = self._get_page_image_path()
        if img_path:
            self._show_image_preview(img_path)

    def _show_image_preview(self, path: str):
        Image, ImageTk, ImageDraw = import_pil()
        if Image is None or ImageDraw is None:
            return
        try:
            img = Image.open(path).convert('RGB')
            
            # Draw table/text regions if in table mode
            if self.preview_mode == 'table' and self.current_page in self.table_results:
                draw = ImageDraw.Draw(img)
                regions = self.table_results[self.current_page]
                for reg in regions:
                    bbox = reg.get('bbox')
                    if bbox and len(bbox) == 4:
                        color = 'green' if reg.get('type') == 'table' else 'blue'
                        width = max(2, int(img.width * 0.005))
                        draw.rectangle(tuple(bbox), outline=color, width=width)
                        
            # Resize for display
            display_size = (360, 460)
            img.thumbnail(display_size, Image.LANCZOS)
            photo = ImageTk.PhotoImage(img)
            self.preview_label.config(image=photo, text='')
            self.preview_label._photo = photo  # keep reference!
            
            # Cleanup temp PDF page if we generated it
            if path.endswith(f"_page{self.current_page}.png") and path != self.current_file:
                 try: os.remove(path)
                 except: pass
                 
        except Exception as e:
            self.preview_label.config(image='', text=f"Preview error:\n{e}")

    def _update_page_nav(self):
        source = self.table_results if self.preview_mode == 'table' else self.ocr_results
        total = len(source)
        if total == 0:
            self.page_var.set("—")
        else:
            self.page_var.set(f"page {self.current_page} / {total}")

    def _prev_page(self):
        if self.current_page > 1:
            self.current_page -= 1
            self._refresh_text()
            self._update_page_nav()
            self._refresh_preview()

    def _next_page(self):
        source = self.table_results if self.preview_mode == 'table' else self.ocr_results
        if self.current_page < len(source):
            self.current_page += 1
            self._refresh_text()
            self._update_page_nav()
            self._refresh_preview()

    # ── Save ───────────────────────────────────────────────────

    def _save_text(self):
        if self.preview_mode == 'table':
            messagebox.showinfo("Save", "In Table Extraction mode, table saving prompts automatically after completion.\nTo save OCR text, please run standard OCR first.")
            return
            
        if not self.ocr_results:
            messagebox.showwarning("Nothing to save", "Run OCR first.")
            return
            
        if not self.current_file:
            return
            
        default = os.path.splitext(os.path.basename(self.current_file))[0] + "_ocr.txt"
        out_path = filedialog.asksaveasfilename(
            title="Save OCR text",
            defaultextension=".txt",
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")],
            initialfile=default
        )
        if not out_path:
            return
        with open(out_path, 'w', encoding='utf-8') as f:
            for page_num in sorted(self.ocr_results.keys()):
                if len(self.ocr_results) > 1:
                    f.write(f"\n{'='*40}\n  PAGE {page_num}\n{'='*40}\n\n")
                f.write(format_results_text(self.ocr_results[page_num], show_confidence=False))
                f.write("\n")
        self.status_var.set(f"Saved: {out_path}")
        messagebox.showinfo("Saved", f"OCR text saved to:\n{out_path}")

    def run(self):
        self.root.mainloop()


if __name__ == '__main__':
    app = PaddleOCRDemo()
    app.run()
